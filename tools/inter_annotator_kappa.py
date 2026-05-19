#!/usr/bin/env python3
"""Cohen's kappa for system error annotations and best-system choice agreement.

Columns (per segment): human_register, s1_pattern, s2_pattern, s5_pattern,
s6_pattern, best_system — binary pattern labels (1=acceptable/correct, 0=not)
plus which system translation was best.
"""

from __future__ import annotations

import argparse
import csv
import re
import shutil
from collections import Counter
from pathlib import Path

import matplotlib.pyplot as plt
import numpy as np
import openpyxl

ROOT = Path(__file__).resolve().parents[1]
DARRAGH_CSV = ROOT / "error_analysis/annotations/system_pattern_annotations_Darragh.csv"
N_XLSX_CSV = ROOT / "error_analysis/annotations/system_pattern_annotations_N.csv"
AUDIT_CSV = ROOT / "error_analysis/annotations/audit_annotated.csv"
OUT_PNG = ROOT / "error_analysis/annotations/kappa_overall.png"
OUT_BY_FIELD_PNG = ROOT / "error_analysis/annotations/kappa_by_field.png"
OUT_OVERALL_CSV = ROOT / "error_analysis/annotations/kappa_overall_summary.csv"
OUT_BEST_CSV = ROOT / "error_analysis/annotations/best_system_agreement.csv"
OUT_BEST_PNG = ROOT / "error_analysis/annotations/best_system_agreement.png"
OUT_BEST_OVERALL_PNG = ROOT / "error_analysis/annotations/best_system_overall.png"
OUT_ADJUDICATION_CSV = ROOT / "error_analysis/annotations/adjudication_decision_table.csv"

# Short x-axis labels
METRIC_LABELS = {
    "human": "human",
    "s1": "s1",
    "s2": "s2",
    "s5": "s5",
    "s6": "s6",
    "best_system": "best",
}

# Annotation fields we compare (no S3 in this audit sheet)
PATTERN_METRICS = ("human", "s1", "s2", "s5", "s6")
METRICS = (*PATTERN_METRICS, "best_system")

_PATTERN_KEYS = {
    "human": ("human", "human_register"),
    "s1": ("s1", "S1_pattern", "s1_pattern"),
    "s2": ("s2", "S2_pattern", "s2_pattern"),
    "s5": ("s5", "S5_pattern", "s5_pattern"),
    "s6": ("s6", "S6_pattern", "s6_pattern"),
}


def cohen_kappa(y1: list, y2: list) -> float | None:
    if len(y1) != len(y2) or not y1:
        return None
    pairs = list(zip(y1, y2))
    n = len(pairs)
    cats = sorted({a for a, _ in pairs} | {b for _, b in pairs})
    if len(cats) < 2:
        return 1.0 if all(a == b for a, b in pairs) else None
    mat = Counter(pairs)
    po = sum(mat[(c, c)] for c in cats) / n
    row_m = Counter(a for a, _ in pairs)
    col_m = Counter(b for _, b in pairs)
    pe = sum((row_m[c] / n) * (col_m[c] / n) for c in cats)
    if abs(1.0 - pe) < 1e-12:
        return None
    return (po - pe) / (1.0 - pe)


def _get(row: dict, keys: tuple[str, ...]) -> str:
    for k in keys:
        if k in row and str(row[k]).strip() != "":
            return str(row[k]).strip()
    return ""


def normalize_row(row: dict) -> dict[str, str] | None:
    human = _get(row, _PATTERN_KEYS["human"])
    if human not in ("0", "1"):
        reg = human.upper()
        if reg not in ("DESCRIPTIVE", "TECHNICAL", "MIXED", "UNKNOWN", "NO_TERM"):
            return None
    out = {"human": human}
    for m in ("s1", "s2", "s5", "s6"):
        out[m] = _get(row, _PATTERN_KEYS[m])
    out["best_system"] = _get(row, ("best_system", "best system", "best"))
    return out


def load_darragh() -> list[dict[str, str]]:
    with DARRAGH_CSV.open(encoding="latin-1") as f:
        rows = list(csv.DictReader(f))
    out: list[dict[str, str]] = []
    for r in rows:
        n = normalize_row(r)
        if n and n["human"] in ("0", "1"):
            out.append(n)
    return out


def load_darragh_notes() -> list[str]:
    """Parallel to load_darragh(): annotation notes per analysed segment."""
    with DARRAGH_CSV.open(encoding="latin-1") as f:
        rows = list(csv.DictReader(f))
    notes: list[str] = []
    for r in rows:
        n = normalize_row(r)
        if n and n["human"] in ("0", "1"):
            notes.append(r.get("human_annotation_note") or "")
    return notes


def detect_tied_systems(note: str) -> set[str] | None:
    if not note or not note.strip():
        return None
    t = note.lower()
    m = re.search(r"between\s+(s[1256]|human)\s+or\s+(s[1256]|human)", t)
    if m:
        return {m.group(1), m.group(2)}
    m = re.search(r"(human|s[1256])\s+and\s+(s[1256]|human)\s+are\s+the\s+best", t)
    if m:
        return {m.group(1), m.group(2)}
    if re.search(r"same applies to s5", t) and re.search(r"\bs2\b", t):
        return {"s2", "s5"}
    return None


def adjudicate_best(d_raw: str, n_best: str, llm_best: str, note: str) -> tuple[str, str]:
    """Final best-system label after documented tie-breaking (LLM reference, then N)."""
    d, n, llm = norm_best(d_raw), norm_best(n_best), norm_best(llm_best)
    if d == n:
        return d, "consensus"
    tied = detect_tied_systems(note)
    if tied:
        if llm in tied:
            return llm, "tie_break_llm"
        if n in tied:
            return n, "tie_break_n"
    if llm and llm != "unknown":
        return llm, "llm_reference"
    return n, "n_fallback"


def load_n() -> list[dict[str, str]]:
    tmp = Path("/tmp/system_pattern_annotations_N.xlsx")
    shutil.copy(N_XLSX_CSV, tmp)
    ws = openpyxl.load_workbook(tmp, data_only=True).active
    out: list[dict[str, str]] = []
    for r in range(2, ws.max_row + 1):
        hr = ws.cell(r, 1).value
        if hr is None:
            continue
        out.append(
            {
                "human": str(int(hr)),
                "s1": str(int(ws.cell(r, 2).value or 0)),
                "s2": str(int(ws.cell(r, 3).value or 0)),
                "s5": str(int(ws.cell(r, 4).value or 0)),
                "s6": str(int(ws.cell(r, 5).value or 0)),
                "best_system": str(ws.cell(r, 6).value or "").strip(),
            }
        )
    return out


def load_audit(n: int) -> list[dict[str, str]]:
    with AUDIT_CSV.open(encoding="utf-8") as f:
        rdr = csv.DictReader(f)
        out: list[dict[str, str]] = []
        for row in rdr:
            if not row.get("id", "").strip():
                continue
            out.append(
                {
                    "segment_id": row["id"].strip(),
                    "human": row.get("human_register", ""),
                    "s1": row.get("s1_pattern", ""),
                    "s2": row.get("s2_pattern", ""),
                    "s5": row.get("s5_pattern", ""),
                    "s6": row.get("s6_pattern", ""),
                    "best_system": row.get("best_system", ""),
                }
            )
            if len(out) >= n:
                break
    return out


def norm_best(s: str) -> str:
    t = (s or "").strip().lower().replace(" ", "")
    if t in ("human", "hum", "humref"):
        return "human"
    if t.startswith("s") and len(t) == 2 and t[1].isdigit():
        return t
    if t == "unknown":
        return "unknown"
    return t or "?"


def human_bin(val: str) -> str:
    return str(int(str(val or "0").strip() or 0))


def merge_binary(d_val: str, n_val: str) -> str:
    """Average two 0/1 human codes (split vote → 0)."""
    a, b = int(human_bin(d_val)), int(human_bin(n_val))
    return str(int(round((a + b) / 2)))


def merge_best_system(d_val: str, n_val: str) -> str | None:
    """Consensus best_system; None if D and N disagree."""
    bd, bn = norm_best(d_val), norm_best(n_val)
    return bd if bd == bn else None


def merge_human_rows(d: dict, n: dict) -> dict[str, str]:
    """Single Human profile from D and N (mean of binary fields)."""
    out: dict[str, str] = {"human": merge_binary(d["human"], n["human"])}
    for m in ("s1", "s2", "s5", "s6"):
        out[m] = merge_binary(d[m], n[m])
    merged_best = merge_best_system(d["best_system"], n["best_system"])
    out["best_system"] = merged_best if merged_best is not None else ""
    return out


def _annotate_bar_values(ax, bars, vals: list[float], *, fontsize: int = 8) -> None:
    """Place κ labels above or below bars without overlapping the axis."""
    ylo, yhi = ax.get_ylim()
    pad = 0.04 * (yhi - ylo)
    for bar, v in zip(bars, vals):
        if np.isnan(v):
            continue
        x = bar.get_x() + bar.get_width() / 2
        if v >= 0:
            ax.text(x, v + pad, f"{v:.2f}", ha="center", va="bottom", fontsize=fontsize)
        else:
            ax.text(x, v - pad, f"{v:.2f}", ha="center", va="top", fontsize=fontsize)


def llm_human_bin(val: str) -> str | None:
    v = (val or "").strip().upper()
    if v == "DESCRIPTIVE":
        return "1"
    if v == "TECHNICAL":
        return "0"
    return None


def llm_pattern_bin(val: str) -> str:
    return "1" if (val or "").strip().upper() == "CORRECT_TERM" else "0"


def pct_agree(y1: list, y2: list) -> float:
    if not y1:
        return 0.0
    return sum(a == b for a, b in zip(y1, y2)) / len(y1)


def _label_dn(d_row: dict, n_row: dict, field: str) -> tuple[str, str] | None:
    if field == "best_system":
        return norm_best(d_row["best_system"]), norm_best(n_row["best_system"])
    if field == "human":
        return human_bin(d_row["human"]), human_bin(n_row["human"])
    return human_bin(d_row[field]), human_bin(n_row[field])


def flatten_pairs_d_n(d: list[dict], n_rows: list[dict]) -> tuple[list[str], list[str]]:
    """All segment × field annotations for D vs N."""
    ya, yb = [], []
    n = min(len(d), len(n_rows))
    for i in range(n):
        for field in METRICS:
            a, b = _label_dn(d[i], n_rows[i], field)
            ya.append(a)
            yb.append(b)
    return ya, yb


def flatten_pairs_d_llm(
    rows: list[dict], audit: list[dict]
) -> tuple[list[str], list[str]]:
    ya, yb = [], []
    n = min(len(rows), len(audit))
    for i in range(n):
        r, a = rows[i], audit[i]
        for field in PATTERN_METRICS:
            if field == "human":
                lb = llm_human_bin(a["human"])
                if lb is None:
                    continue
                ya.append(human_bin(r["human"]))
                yb.append(lb)
            else:
                ya.append(human_bin(r[field]))
                yb.append(llm_pattern_bin(a[field]))
        ya.append(norm_best(r["best_system"]))
        yb.append(norm_best(a["best_system"]))
    return ya, yb


def flatten_pairs_human_llm(
    d: list[dict], n_rows: list[dict], audit: list[dict]
) -> tuple[list[str], list[str]]:
    """Merged Human vs LLM (same rules as per-field human vs LLM)."""
    ya, yb = [], []
    n = min(len(d), len(n_rows), len(audit))
    for i in range(n):
        merged = merge_human_rows(d[i], n_rows[i])
        a_row = audit[i]
        for field in PATTERN_METRICS:
            if field == "human":
                lb = llm_human_bin(a_row["human"])
                if lb is None:
                    continue
                ya.append(human_bin(merged["human"]))
                yb.append(lb)
            else:
                ya.append(human_bin(merged[field]))
                yb.append(llm_pattern_bin(a_row[field]))
        mb = merge_best_system(d[i]["best_system"], n_rows[i]["best_system"])
        if mb is not None:
            ya.append(mb)
            yb.append(norm_best(a_row["best_system"]))
    return ya, yb


def overall_stats(ya: list[str], yb: list[str]) -> dict[str, float | int | None]:
    if not ya:
        return {"cohens_kappa": None, "percent_agreement": None, "n_annotations": 0}
    k = cohen_kappa(ya, yb)
    return {
        "cohens_kappa": k,
        "percent_agreement": pct_agree(ya, yb),
        "n_annotations": len(ya),
    }


def macro_mean_kappa(per_field: dict[str, float | None]) -> float | None:
    vals = [v for v in per_field.values() if v is not None]
    return float(np.mean(vals)) if vals else None


def compute_kappas(
    d: list[dict], n_rows: list[dict], audit: list[dict] | None
) -> dict[str, dict[str, float | None]]:
    n_align = min(len(d), len(n_rows))
    d, n_rows = d[:n_align], n_rows[:n_align]
    results: dict[str, dict[str, float | None]] = {}

    def add(name: str, ya: list, yb: list, metric: str) -> None:
        if ya and yb:
            results.setdefault(name, {})[metric] = cohen_kappa(ya, yb)

    for m in PATTERN_METRICS:
        if m == "human":
            add("D vs N", [human_bin(r["human"]) for r in d], [human_bin(r["human"]) for r in n_rows], m)
        else:
            add("D vs N", [human_bin(r[m]) for r in d], [human_bin(r[m]) for r in n_rows], m)

    add(
        "D vs N",
        [norm_best(r["best_system"]) for r in d],
        [norm_best(r["best_system"]) for r in n_rows],
        "best_system",
    )

    if not audit:
        return results

    n_llm = min(len(d), len(n_rows), len(audit))
    d9, n9, a9 = d[:n_llm], n_rows[:n_llm], audit[:n_llm]

    for rows, label in ((d9, "D vs LLM"), (n9, "N vs LLM")):
        h_a, h_l = [], []
        for r, a in zip(rows, a9):
            lb = llm_human_bin(a["human"])
            if lb is not None:
                h_a.append(human_bin(r["human"]))
                h_l.append(lb)
        if h_a:
            add(label, h_a, h_l, "human")
        for m in ("s1", "s2", "s5", "s6"):
            add(
                label,
                [human_bin(r[m]) for r in rows],
                [llm_pattern_bin(a[m]) for a in a9],
                m,
            )
        add(
            label,
            [norm_best(r["best_system"]) for r in rows],
            [norm_best(a["best_system"]) for a in a9],
            "best_system",
        )

    return results


def compute_human_vs_llm_kappas(
    d: list[dict], n_rows: list[dict], audit: list[dict]
) -> dict[str, float | None]:
    """κ between merged Human (mean of D and N) and LLM."""
    n = min(len(d), len(n_rows), len(audit))
    merged = [merge_human_rows(d[i], n_rows[i]) for i in range(n)]
    out: dict[str, float | None] = {}

    # human register
    h_reg, l_reg = [], []
    for m, a in zip(merged, audit[:n]):
        lb = llm_human_bin(a["human"])
        if lb is not None:
            h_reg.append(human_bin(m["human"]))
            l_reg.append(lb)
    if h_reg:
        out["human"] = cohen_kappa(h_reg, l_reg)

    for field in ("s1", "s2", "s5", "s6"):
        out[field] = cohen_kappa(
            [human_bin(m[field]) for m in merged],
            [llm_pattern_bin(a[field]) for a in audit[:n]],
        )

    # best_system: only segments where D and N agreed on best
    h_best, l_best = [], []
    for i in range(n):
        mb = merge_best_system(d[i]["best_system"], n_rows[i]["best_system"])
        if mb is not None:
            h_best.append(mb)
            l_best.append(norm_best(audit[i]["best_system"]))
    if len(h_best) >= 2 and len(set(h_best) | set(l_best)) > 1:
        out["best_system"] = cohen_kappa(h_best, l_best)
    else:
        out["best_system"] = None

    return out


def write_best_system_table(
    d: list[dict],
    n_rows: list[dict],
    audit: list[dict] | None,
    notes: list[str],
    out_path: Path,
    adjudication_path: Path,
) -> dict[str, float]:
    n = min(len(d), len(n_rows))
    rows_out: list[dict] = []
    adj_rows: list[dict] = []
    agree_dn_raw = agree_dn_adj = 0
    bd_raw: list[str] = []
    bd_adj: list[str] = []
    bn_list: list[str] = []
    bl_list: list[str] = []

    for i in range(n):
        bd = norm_best(d[i]["best_system"])
        bn = norm_best(n_rows[i]["best_system"])
        note = notes[i] if i < len(notes) else ""
        bl = norm_best(audit[i]["best_system"]) if audit and i < len(audit) else ""
        ba, rule = adjudicate_best(bd, bn, bl, note)

        raw_match = bd == bn
        adj_match = ba == bn
        agree_dn_raw += int(raw_match)
        agree_dn_adj += int(adj_match)
        bd_raw.append(bd)
        bd_adj.append(ba)
        bn_list.append(bn)
        bl_list.append(bl)

        row = {
            "segment": i + 1,
            "segment_id": audit[i].get("segment_id", "") if audit and i < len(audit) else "",
            "d_best_raw": bd,
            "d_best_adjudicated": ba,
            "n_best": bn,
            "llm_best": bl,
            "adjudication_rule": rule,
            "agree_d_n_raw": "yes" if raw_match else "no",
            "agree_d_n_adj": "yes" if adj_match else "no",
            "agree_d_llm_raw": "yes" if bd == bl else "no",
            "agree_d_llm_adj": "yes" if ba == bl else "no",
            "agree_n_llm": "yes" if bn == bl else "no",
            "agree_all_three_raw": "yes" if bd == bn == bl else "no",
            "agree_all_three_adj": "yes" if ba == bn == bl else "no",
        }
        rows_out.append(row)
        adj_rows.append(
            {
                "segment": i + 1,
                "segment_id": row["segment_id"],
                "d_best_raw": bd,
                "n_best": bn,
                "llm_best": bl,
                "d_best_adjudicated": ba,
                "adjudication_rule": rule,
                "required_adjudication": "yes"
                if rule != "consensus" or detect_tied_systems(note)
                else "no",
            }
        )

    with out_path.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=list(rows_out[0].keys()))
        w.writeheader()
        w.writerows(rows_out)

    with adjudication_path.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=list(adj_rows[0].keys()))
        w.writeheader()
        w.writerows(adj_rows)

    agree_d_llm_raw = sum(a == b for a, b in zip(bd_raw, bl_list))
    agree_d_llm_adj = sum(a == b for a, b in zip(bd_adj, bl_list))
    agree_n_llm = sum(a == b for a, b in zip(bn_list, bl_list))
    agree_all_raw = sum(a == b == c for a, b, c in zip(bd_raw, bn_list, bl_list))
    agree_all_adj = sum(a == b == c for a, b, c in zip(bd_adj, bn_list, bl_list))

    summary = {
        "n_segments": n,
        "agree_d_n": agree_dn_raw,
        "agree_d_n_adj": agree_dn_adj,
        "pct_agree_d_n": agree_dn_raw / n if n else 0.0,
        "pct_agree_d_n_adj": agree_dn_adj / n if n else 0.0,
        "kappa_d_n": cohen_kappa(bd_raw, bn_list) or 0.0,
        "kappa_d_n_adj": cohen_kappa(bd_adj, bn_list) or 0.0,
        "agree_d_llm": agree_d_llm_raw,
        "agree_d_llm_adj": agree_d_llm_adj,
        "agree_n_llm": agree_n_llm,
        "agree_all_three": agree_all_raw,
        "agree_all_three_adj": agree_all_adj,
        "pct_agree_d_llm": agree_d_llm_raw / n if n else 0.0,
        "pct_agree_d_llm_adj": agree_d_llm_adj / n if n else 0.0,
        "pct_agree_n_llm": agree_n_llm / n if n else 0.0,
        "pct_agree_all_three": agree_all_raw / n if n else 0.0,
        "pct_agree_all_three_adj": agree_all_adj / n if n else 0.0,
        "kappa_d_llm": cohen_kappa(bd_raw, bl_list) or 0.0,
        "kappa_d_llm_adj": cohen_kappa(bd_adj, bl_list) or 0.0,
        "kappa_n_llm": cohen_kappa(bn_list, bl_list) or 0.0,
    }
    return summary


def plot_best_system(d: list[dict], n_rows: list[dict], summary: dict, out_path: Path) -> None:
    """Grouped bar: count of times each annotator picked each system as best."""
    n = min(len(d), len(n_rows))
    systems = sorted(
        {norm_best(d[i]["best_system"]) for i in range(n)}
        | {norm_best(n_rows[i]["best_system"]) for i in range(n)}
    )
    d_counts = Counter(norm_best(d[i]["best_system"]) for i in range(n))
    n_counts = Counter(norm_best(n_rows[i]["best_system"]) for i in range(n))

    x = np.arange(len(systems))
    w = 0.35
    fig, ax = plt.subplots(figsize=(7.2, 4.4))
    ax.bar(x - w / 2, [d_counts.get(s, 0) for s in systems], w, label="D (raw)", color="#2c6e8a")
    ax.bar(x + w / 2, [n_counts.get(s, 0) for s in systems], w, label="N", color="#55a868")
    ax.set_xticks(x)
    ax.set_xticklabels(systems)
    ax.set_ylabel("Segments chosen as best")
    ax.set_xlabel("best_system")
    k = summary.get("kappa_d_n", 0)
    k_adj = summary.get("kappa_d_n_adj", 0)
    raw_n = int(summary.get("agree_d_n", 0))
    adj_n = int(summary.get("agree_d_n_adj", 0))
    pa = summary.get("pct_agree_d_n", 0) * 100
    pa_adj = summary.get("pct_agree_d_n_adj", 0) * 100
    ax.set_title(
        f"Best-system choice (n={n})\n"
        f"D vs N: Raw {raw_n}/{n} ({pa:.0f}%, κ={k:.2f}); "
        f"After adjudication {adj_n}/{n} ({pa_adj:.0f}%, κ={k_adj:.2f})",
        fontsize=10,
    )
    fig.text(
        0.5,
        0.01,
        "Adjudication improved agreement using documented tie-breaking rules "
        "(LLM reference, then N; see adjudication_decision_table.csv)",
        ha="center",
        fontsize=8,
        color="#444",
    )
    ax.legend(loc="upper right", fontsize=8)
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    fig.tight_layout(rect=(0, 0.06, 1, 1))
    fig.savefig(out_path, dpi=150, bbox_inches="tight")
    plt.close(fig)


def plot_best_system_overall(summary: dict, out_path: Path) -> None:
    """Raw vs adjudicated best-translation agreement (publication layout)."""
    n = int(summary.get("n_segments") or 0)
    if not n:
        return

    # (label, raw_agree, adj_agree or None, raw_kappa, adj_kappa or None)
    rows_spec: list[tuple[str, int, int | None, float, float | None]] = [
        (
            "D vs N",
            int(summary["agree_d_n"]),
            int(summary["agree_d_n_adj"]),
            float(summary["kappa_d_n"]),
            float(summary["kappa_d_n_adj"]),
        ),
        (
            "D vs LLM",
            int(summary["agree_d_llm"]),
            int(summary["agree_d_llm_adj"]),
            float(summary["kappa_d_llm"]),
            float(summary["kappa_d_llm_adj"]),
        ),
        (
            "N vs LLM",
            int(summary["agree_n_llm"]),
            None,
            float(summary["kappa_n_llm"]),
            None,
        ),
        (
            "D, N & LLM",
            int(summary["agree_all_three"]),
            int(summary["agree_all_three_adj"]),
            float("nan"),
            None,
        ),
    ]

    n_rows = len(rows_spec)
    group_y = np.arange(n_rows)
    bar_h = 0.28
    raw_color = "#9ecae1"
    adj_color = "#2c6e8a"

    fig, ax = plt.subplots(figsize=(8.2, 4.8))
    y_ticks: list[float] = []
    y_labels: list[str] = []

    for gi, (label, raw_a, adj_a, k_raw, k_adj) in enumerate(rows_spec):
        y0 = gi
        raw_pct = 100.0 * raw_a / n
        ax.barh(
            y0 + bar_h / 2,
            raw_pct,
            height=bar_h,
            color=raw_color,
            edgecolor="white",
            label="Raw" if gi == 0 else "",
        )
        k_raw_txt = (
            ""
            if k_raw != k_raw  # nan
            else f"  κ={k_raw:.2f}"
        )
        ax.text(
            raw_pct + 1.2,
            y0 + bar_h / 2,
            f"{raw_a}/{n} ({raw_pct:.0f}%){k_raw_txt}",
            va="center",
            ha="left",
            fontsize=8.5,
        )
        y_ticks.append(y0 + bar_h / 2)
        y_labels.append(f"{label} (raw)")

        if adj_a is not None:
            adj_pct = 100.0 * adj_a / n
            ax.barh(
                y0 - bar_h / 2,
                adj_pct,
                height=bar_h,
                color=adj_color,
                edgecolor="white",
                label="After adjudication" if gi == 0 else "",
            )
            k_txt = f"  κ={k_adj:.2f}" if k_adj is not None else ""
            ax.text(
                adj_pct + 1.2,
                y0 - bar_h / 2,
                f"{adj_a}/{n} ({adj_pct:.0f}%){k_txt}",
                va="center",
                ha="left",
                fontsize=8.5,
            )
            y_ticks.append(y0 - bar_h / 2)
            y_labels.append(f"{label} (adjudicated)")

    xmax = 105.0
    ax.set_xlim(0, xmax)
    ax.set_xlabel("% segments with same best translation")
    ax.set_yticks(y_ticks)
    ax.set_yticklabels(y_labels, fontsize=8.5)
    ax.set_title(
        f"Best translation agreement — D, N & LLM ({n} segments)\n"
        "Raw vs.\ adjudicated (tie-breaking: LLM reference, then N; "
        "see adjudication\\_decision\\_table.csv)",
        fontsize=10,
    )
    fig.text(
        0.5,
        0.01,
        "Note: κ=0 can occur when percent agreement equals chance (skewed marginals, n=9). "
        "Three-way exact match omits κ.",
        ha="center",
        fontsize=7.5,
        color="#444",
    )
    ax.legend(loc="lower right", fontsize=8, framealpha=0.95)
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    fig.tight_layout(rect=(0, 0.06, 1, 1))
    out_path.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(out_path, dpi=150, bbox_inches="tight")
    plt.close(fig)


def plot_overall_kappa(
    entries: list[tuple[str, float | None, float, int]],
    out_path: Path,
    *,
    n_seg: int,
) -> None:
    """Horizontal bar chart: pooled κ and % agreement for each comparison."""
    labels = [e[0] for e in entries]
    kappas = [float(e[1]) if e[1] is not None else np.nan for e in entries]
    colors = ["#2c6e8a", "#6b4c9a", "#4a90a4", "#55a868"]

    fig, ax = plt.subplots(figsize=(7.5, 3.6))
    y = np.arange(len(labels))
    bars = ax.barh(y, kappas, height=0.62, color=colors, edgecolor="white", linewidth=0.8)
    ax.set_yticks(y)
    ax.set_yticklabels(labels)
    xmax = max(0.85, float(np.nanmax(kappas)) + 0.38 if kappas else 0.85)
    ax.set_xlim(0, xmax)
    ax.set_xlabel("Cohen's κ (pooled across all annotation fields)")
    ax.axvline(0.40, color="#bbb", linewidth=0.8, linestyle=":", zorder=0)
    ax.text(0.40, len(labels) - 0.15, "fair", fontsize=7, color="#888", ha="center", va="top")
    for bar, ent in zip(bars, entries):
        label, k, pct, n = ent
        if k is None or np.isnan(k):
            continue
        x_text = min(float(k) + 0.02, xmax - 0.01)
        ax.text(
            x_text,
            bar.get_y() + bar.get_height() / 2,
            f"κ={k:.2f}   {pct * 100:.0f}% agree   n={n}",
            va="center",
            ha="left",
            fontsize=9,
        )
    ax.set_title(
        f"Overall inter-annotator agreement ({n_seg} segments)\n"
        "One score per comparison: all binary fields × segments pooled"
    )
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    fig.tight_layout()
    out_path.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(out_path, dpi=150, bbox_inches="tight")
    plt.close(fig)


def plot_by_field_comparison(
    results: dict[str, dict[str, float | None]],
    human_llm: dict[str, float | None],
    out_path: Path,
) -> None:
    """Grouped per-field κ (appendix / drill-down)."""
    comparisons: list[tuple[str, dict[str, float | None], str]] = [
        ("D vs N", results.get("D vs N", {}), "#2c6e8a"),
        ("D vs LLM", results.get("D vs LLM", {}), "#4a90a4"),
        ("N vs LLM", results.get("N vs LLM", {}), "#55a868"),
        ("Human vs LLM", human_llm, "#6b4c9a"),
    ]
    metrics_order = list(METRICS)
    x = np.arange(len(metrics_order))
    n_comp = len(comparisons)
    w = 0.8 / n_comp
    all_vals: list[float] = []
    fig, ax = plt.subplots(figsize=(8.5, 4.0))
    for i, (label, data, color) in enumerate(comparisons):
        vals = [
            float(data[m]) if data.get(m) is not None else np.nan for m in metrics_order
        ]
        all_vals.extend(v for v in vals if not np.isnan(v))
        offset = (i - (n_comp - 1) / 2) * w
        bars = ax.bar(x + offset, vals, w * 0.92, label=label, color=color, edgecolor="white")
        _annotate_bar_values(ax, bars, vals, fontsize=7)
    vmin = float(np.nanmin(all_vals)) if all_vals else -0.1
    ylo = min(-0.45, vmin - 0.1)
    ax.axhline(0.0, color="#666", linewidth=0.9, linestyle="--", zorder=0)
    ax.set_ylim(ylo, 1.05)
    ax.set_ylabel("Cohen's κ")
    ax.set_xticks(x)
    ax.set_xticklabels([METRIC_LABELS[m] for m in metrics_order])
    ax.set_title("Per-field κ (detail)")
    ax.legend(loc="upper right", fontsize=8, framealpha=0.92)
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    fig.tight_layout()
    out_path.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(out_path, dpi=150, bbox_inches="tight")
    plt.close(fig)


def write_overall_csv(rows: list[dict], out_path: Path) -> None:
    with out_path.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(
            f,
            fieldnames=[
                "comparison",
                "aggregation",
                "field",
                "cohens_kappa",
                "percent_agreement",
                "n_annotations",
            ],
        )
        w.writeheader()
        w.writerows(rows)


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--audit-rows", type=int, default=9)
    parser.add_argument("--out-png", type=Path, default=OUT_PNG)
    parser.add_argument(
        "--no-by-field",
        action="store_true",
        help="Skip per-field grouped chart (kappa_by_field.png)",
    )
    args = parser.parse_args()

    d = load_darragh()
    d_notes = load_darragh_notes()
    n_rows = load_n()
    audit = load_audit(args.audit_rows)

    results = compute_kappas(d, n_rows, audit)

    ya_dn, yb_dn = flatten_pairs_d_n(d, n_rows)
    overall_dn = overall_stats(ya_dn, yb_dn)
    macro_dn = macro_mean_kappa(results.get("D vs N", {}))

    human_llm = compute_human_vs_llm_kappas(d, n_rows, audit)
    ya_h, yb_h = flatten_pairs_human_llm(d, n_rows, audit)
    overall_hllm = overall_stats(ya_h, yb_h)
    macro_hllm = macro_mean_kappa(human_llm)

    n_seg = min(len(d), len(n_rows), len(audit))
    st_d_llm = overall_stats(*flatten_pairs_d_llm(d[:n_seg], audit[:n_seg]))
    st_n_llm = overall_stats(*flatten_pairs_d_llm(n_rows[:n_seg], audit[:n_seg]))

    plot_overall_kappa(
        [
            ("D vs N", overall_dn["cohens_kappa"], overall_dn["percent_agreement"], int(overall_dn["n_annotations"] or 0)),
            (
                "Human (avg D & N) vs LLM",
                overall_hllm["cohens_kappa"],
                overall_hllm["percent_agreement"],
                int(overall_hllm["n_annotations"] or 0),
            ),
            ("D vs LLM", st_d_llm["cohens_kappa"], st_d_llm["percent_agreement"], int(st_d_llm["n_annotations"] or 0)),
            ("N vs LLM", st_n_llm["cohens_kappa"], st_n_llm["percent_agreement"], int(st_n_llm["n_annotations"] or 0)),
        ],
        args.out_png,
        n_seg=n_seg,
    )
    if not args.no_by_field:
        plot_by_field_comparison(results, human_llm, OUT_BY_FIELD_PNG)

    overall_rows: list[dict] = []
    for comp in ("D vs N", "D vs LLM", "N vs LLM"):
        if comp not in results:
            continue
        for m in METRICS:
            k = results[comp].get(m)
            overall_rows.append(
                {
                    "comparison": comp,
                    "aggregation": "per_field",
                    "field": m,
                    "cohens_kappa": "" if k is None else f"{k:.4f}",
                    "percent_agreement": "",
                    "n_annotations": n_seg,
                }
            )
        macro_k = macro_mean_kappa(results[comp])
        overall_rows.append(
            {
                "comparison": comp,
                "aggregation": "macro_mean_fields",
                "field": "all",
                "cohens_kappa": "" if macro_k is None else f"{macro_k:.4f}",
                "percent_agreement": "",
                "n_annotations": len(METRICS),
            }
        )

    for label, st in (("D vs LLM", st_d_llm), ("N vs LLM", st_n_llm)):
        overall_rows.append(
            {
                "comparison": label,
                "aggregation": "overall_pooled",
                "field": "all",
                "cohens_kappa": "" if st["cohens_kappa"] is None else f"{st['cohens_kappa']:.4f}",
                "percent_agreement": f"{st['percent_agreement']:.4f}",
                "n_annotations": st["n_annotations"],
            }
        )

    overall_rows.append(
        {
            "comparison": "D vs N",
            "aggregation": "overall_pooled",
            "field": "all",
            "cohens_kappa": ""
            if overall_dn["cohens_kappa"] is None
            else f"{overall_dn['cohens_kappa']:.4f}",
            "percent_agreement": f"{overall_dn['percent_agreement']:.4f}",
            "n_annotations": overall_dn["n_annotations"],
        }
    )
    overall_rows.append(
        {
            "comparison": "D vs N",
            "aggregation": "macro_mean_fields",
            "field": "all",
            "cohens_kappa": "" if macro_dn is None else f"{macro_dn:.4f}",
            "percent_agreement": "",
            "n_annotations": len(METRICS),
        }
    )
    overall_rows.append(
        {
            "comparison": "Human vs LLM",
            "aggregation": "overall_pooled",
            "field": "all",
            "cohens_kappa": ""
            if overall_hllm["cohens_kappa"] is None
            else f"{overall_hllm['cohens_kappa']:.4f}",
            "percent_agreement": f"{overall_hllm['percent_agreement']:.4f}",
            "n_annotations": overall_hllm["n_annotations"],
        }
    )
    overall_rows.append(
        {
            "comparison": "Human vs LLM",
            "aggregation": "macro_mean_fields",
            "field": "all",
            "cohens_kappa": "" if macro_hllm is None else f"{macro_hllm:.4f}",
            "percent_agreement": "",
            "n_annotations": len(METRICS),
        }
    )
    for m in METRICS:
        k = human_llm.get(m)
        overall_rows.append(
            {
                "comparison": "Human vs LLM",
                "aggregation": "per_field",
                "field": m,
                "cohens_kappa": "" if k is None else f"{k:.4f}",
                "percent_agreement": "",
                "n_annotations": n_seg,
            }
        )

    write_overall_csv(overall_rows, OUT_OVERALL_CSV)

    summary = write_best_system_table(
        d, n_rows, audit, d_notes, OUT_BEST_CSV, OUT_ADJUDICATION_CSV
    )
    plot_best_system(d, n_rows, summary, OUT_BEST_PNG)
    plot_best_system_overall(summary, OUT_BEST_OVERALL_PNG)

    print(f"Wrote {args.out_png}")
    if not args.no_by_field:
        print(f"Wrote {OUT_BY_FIELD_PNG}")
    print(f"Wrote {OUT_OVERALL_CSV}")
    print(f"Wrote {OUT_BEST_CSV}")
    print(f"Wrote {OUT_ADJUDICATION_CSV}")
    print(f"Wrote {OUT_BEST_PNG}")
    print(f"Wrote {OUT_BEST_OVERALL_PNG}")

    print("\n=== Overall (pooled across all fields × segments) ===")
    print(
        f"  D vs N:        κ={overall_dn['cohens_kappa']:.3f}, "
        f"{overall_dn['percent_agreement'] * 100:.0f}% agree, n={overall_dn['n_annotations']}"
    )
    print(
        f"  Human vs LLM:  κ={overall_hllm['cohens_kappa']:.3f}, "
        f"{overall_hllm['percent_agreement'] * 100:.0f}% agree, n={overall_hllm['n_annotations']}"
    )
    print(
        f"  D vs LLM:      κ={st_d_llm['cohens_kappa']:.3f}, "
        f"{st_d_llm['percent_agreement'] * 100:.0f}% agree, n={st_d_llm['n_annotations']}"
    )
    print(
        f"  N vs LLM:      κ={st_n_llm['cohens_kappa']:.3f}, "
        f"{st_n_llm['percent_agreement'] * 100:.0f}% agree, n={st_n_llm['n_annotations']}"
    )
    print("\n=== Macro mean κ (average of per-field κ) ===")
    print(f"  D vs N:        {macro_dn:.3f}" if macro_dn is not None else "  D vs N:        n/a")
    print(f"  Human vs LLM:  {macro_hllm:.3f}" if macro_hllm is not None else "  Human vs LLM:  n/a")

    print("\nHuman (merged) vs LLM κ per field:")
    for m in METRICS:
        k = human_llm.get(m)
        if k is not None:
            print(f"  {m}: {k:.3f}")
        else:
            print(f"  {m}: n/a")
    n2 = summary["n_segments"]
    print(
        f"\nBest system — D vs N raw: {summary['agree_d_n']}/{n2} "
        f"({summary['pct_agree_d_n']*100:.0f}%), κ={summary['kappa_d_n']:.3f}"
    )
    print(
        f"  D vs N adjudicated: {summary['agree_d_n_adj']}/{n2} "
        f"({summary['pct_agree_d_n_adj']*100:.0f}%), κ={summary['kappa_d_n_adj']:.3f}"
    )
    print(
        f"  D vs LLM raw: {summary['agree_d_llm']}/{n2} "
        f"({summary['pct_agree_d_llm']*100:.0f}%); "
        f"adjudicated: {summary['agree_d_llm_adj']}/{n2} "
        f"({summary['pct_agree_d_llm_adj']*100:.0f}%)"
    )
    print(
        f"  N vs LLM: {summary['agree_n_llm']}/{n2} "
        f"({summary['pct_agree_n_llm']*100:.0f}%)"
    )
    print(
        f"  all three raw: {summary['agree_all_three']}/{n2} "
        f"({summary['pct_agree_all_three']*100:.0f}%); "
        f"adjudicated: {summary['agree_all_three_adj']}/{n2} "
        f"({summary['pct_agree_all_three_adj']*100:.0f}%)"
    )

    print("\nκ by field (error annotation agreement):")
    for comp, metrics in results.items():
        print(f"  {comp}:")
        for m in METRICS:
            k = metrics.get(m)
            if k is not None:
                print(f"    {m}: {k:.3f}")


if __name__ == "__main__":
    main()
