#!/usr/bin/env python3
"""Resolve tied best_system choices in Darragh annotations using LLM or N."""

from __future__ import annotations

import csv
import os
import re
import shutil
import subprocess
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
DARRAGH_CSV = ROOT / "error_analysis/annotations/system_pattern_annotations_Darragh.csv"
N_XLSX_CSV = ROOT / "error_analysis/annotations/system_pattern_annotations_N.csv"
AUDIT_CSV = ROOT / "error_analysis/annotations/audit_annotated.csv"
KAPPA_SCRIPT = ROOT / "tools/inter_annotator_kappa.py"


def norm_best(s: str) -> str:
    t = (s or "").strip().lower().replace(" ", "")
    if t in ("human", "hum", "humref"):
        return "human"
    if t.startswith("s") and len(t) == 2 and t[1].isdigit():
        return t
    if t == "unknown":
        return "unknown"
    return t or ""


def display_best(s: str) -> str:
    n = norm_best(s)
    return "human" if n == "human" else n


def load_n_best() -> list[str]:
    tmp = Path("/tmp/system_pattern_annotations_N.xlsx")
    shutil.copy(N_XLSX_CSV, tmp)
    ws = openpyxl.load_workbook(tmp, data_only=True).active
    out: list[str] = []
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, 1).value is None:
            continue
        out.append(norm_best(str(ws.cell(r, 6).value or "")))
    return out


def load_llm_best(n: int) -> list[str]:
    with AUDIT_CSV.open(encoding="utf-8") as f:
        rows = list(csv.DictReader(f))[:n]
    return [norm_best(r.get("best_system", "")) for r in rows]


def detect_tied_systems(note: str) -> set[str] | None:
    if not note or not note.strip():
        return None
    t = note.lower()

    m = re.search(r"between\s+(s[1256]|human)\s+or\s+(s[1256]|human)", t)
    if m:
        return {m.group(1), m.group(2)}

    m = re.search(r"(human|s[1256])\s+and\s+(s[1256]|human)\s+are\s+the\s+best", t)
    if m:
        return {m.group(1), m.group(2)}

    if re.search(r"same applies to s5", t) and re.search(r"\bs2\b", t):
        return {"s2", "s5"}

    return None


def resolve_tie(tied: set[str], llm: str, n_best: str) -> str | None:
    if llm in tied:
        return llm
    if n_best in tied:
        return n_best
    if llm and llm != "unknown" and llm == n_best:
        return llm
    return None


def resolve_missing(llm: str, n_best: str) -> str | None:
    if llm and llm != "unknown":
        return llm
    if n_best:
        return n_best
    return None


def main() -> int:
    n_best_list = load_n_best()
    with DARRAGH_CSV.open(encoding="latin-1", newline="") as fh:
        reader = csv.DictReader(fh)
        rows = list(reader)
        fieldnames = reader.fieldnames or list(rows[0].keys())

    llm_best_list = load_llm_best(len(rows))
    changes: list[str] = []

    for i, row in enumerate(rows):
        note = row.get("human_annotation_note") or ""
        current = norm_best(row.get("best_system", ""))
        llm = llm_best_list[i] if i < len(llm_best_list) else ""
        n_best = n_best_list[i] if i < len(n_best_list) else ""
        seg = i + 1
        new_val: str | None = None
        reason = ""

        tied = detect_tied_systems(note)
        if tied:
            pick = resolve_tie(tied, llm, n_best)
            if pick and pick != current:
                new_val = pick
                reason = f"tie {sorted(tied)} → {pick} (LLM={llm}, N={n_best})"
        elif not current:
            pick = resolve_missing(llm, n_best)
            if pick:
                new_val = pick
                reason = f"empty best → {pick} (LLM={llm}, N={n_best})"

        if new_val:
            row["best_system"] = display_best(new_val)
            changes.append(f"  seg {seg}: {current or '(empty)'} → {new_val}  ({reason})")

        b = norm_best(row.get("best_system", ""))
        if b:
            row["best_system"] = display_best(b)

    with DARRAGH_CSV.open("w", encoding="latin-1", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames, lineterminator="\n")
        w.writeheader()
        w.writerows(rows)

    print(f"Updated {DARRAGH_CSV}")
    if changes:
        print("Changes:")
        print("\n".join(changes))
    else:
        print("No best_system changes needed.")

    print("\nRegenerating agreement charts and CSVs …")
    env = {**os.environ, "PYTHONPATH": str(ROOT)}
    subprocess.run([sys.executable, str(KAPPA_SCRIPT)], cwd=ROOT, env=env, check=True)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
